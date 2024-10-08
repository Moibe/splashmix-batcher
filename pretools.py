import os 
import time
import tools
import servidor
import requests
import pandas as pd
import nycklar.nodes as nodes
import configuracion.globales as globales
from openpyxl import Workbook, load_workbook
import configuracion.configuracion as configuracion
from objetosCreacion import Hotgirl, Superhero

def creaDirectorioInicial(sesion): 
    """
    Crear los directorios donde se guardarán las fotos, con un nombre que se recíbe como parámetro.

    Parameters:
    archivo (str): Nombre y extensión del archivo que procesaremos.

    Returns:
    dataframe:Regresa dataframe que se usará a través de los procesos.

    """
      

    # Define the desired directory path
    target_dir = os.path.join('imagenes', 'fuentes', sesion)
    
    if not os.path.exists(target_dir):
        os.makedirs(target_dir)
        print("Directorio creado...")

    print("Directorio existe")
    

def creaExcel(filename):
      
    #Future, poner guardado por interrupción también en creaExcel.
    
    #ESTRUCTURA INICIAL
    #La estructura inicial solo se hará si no existe el arhivo, si no, de lo contrario estaría borrando lo que ya...
    #se hizo.

    #Define si ya existe el archivo de excel o se está completando un proceso previamente iniciado.
    if not os.path.exists(globales.excel_results_path + configuracion.sesion + '.xlsx'):

        print("El archivo de excel no existe, se creará en éste momento.")
        dataframe = pd.read_excel(globales.excel_source_path + filename)        

        #Si no existe, entonces si hacemos toda la estructura inicial.
        #Importante: Crea las nuevas columnas que necesitará:
        #Future, revisa si podría no crearlas, ya vez que actualizaRow las crea al vuelo.
        #Future Importante: Checa al crear un archivo nuevo si agregar File Path afecto en algo.
        #Future: Y checa si crear URL, afecta algo.
        dataframe['Source Path'] = ''
        dataframe['Source URL'] = ''
        dataframe['Name'] = ''
        dataframe['Download Status'] = ''
        dataframe['Take'] = ''
        dataframe['File'] = ''
        dataframe['File Path'] = ''
        dataframe['Diffusion Status'] = ''
        dataframe['URL'] = ''

        #Ve si afecta actualizar el excel antes de entregar el dataframe.
        #IMPORTANTE: Quizá no se necesita hacer ésta escritura pq si hace la escritura final. Prueba.
        print("Guardando la estructura inicial del archivo de excel.")
        tools.df2Excel(dataframe, configuracion.sesion + '.xlsx')
    
    #Y exista o se haya creado apenas, adquiriremos el dataframe de results! y hará la creación de los IDs.
    #Ésto ya tiene control para empezar desde donde ibamos.
    #Future, empezar desde donde nos quedámos asume que el excel siempre estará en orden, porque solo suma...
    #...el index adicional. Para evitar ésto, en el futuro usa index row. Aunque poría usar más procesamiento.
    dataframe = pd.read_excel(globales.excel_results_path + filename) 

    #CREACIÓN DE IDs DE ARCHIVOS.
    
    lote_total = len(dataframe)
    print("El tamaño total del lote es: ", lote_total)

    y = 0
    lote_procesar = 1

    while y < lote_procesar:
        #Future: El while si reanuda, pero para volver a hacer ese ciclo, modificarlo o quitarlo. 

        print("Entrando al while...")
        
        por_procesar = dataframe[dataframe['Name'].isna()]
        print("Ésta es el dataframe 'por procesar'...")
        print(por_procesar)
        lote_procesar = len(por_procesar)
        
        print("Nos faltan por nombrar: ", lote_procesar)

        #El nuevo indicé marcará cuanto sumar al indice para que ponga las celdas en el lugar correcto.
        #Basado en donde empezó.
        nu_index = lote_total - lote_procesar

        #Está bien que haya proceso de reiniciar donde ibamos, porque no solo es que sean lotes muy largos, pero ...
        #el proceso para extraer su nombre podría ser complicado.

        print(f"Por lo tanto llevamos {nu_index} imagenes nombradas.")
        
        columna = por_procesar['Source']   
        print("El resultado de columna es....")
        print(columna)     
        
        #IMPORTANTE: Si obtiene correctamente la columna, pero la i no funcionará ahora, necesitas indexRow...
        #Considera que ese proceso podría ser más lento y que en verdad convenga más crear otra vez todos los ids.

        #Importante: Ésta es la parte que hace el ciclo que genera las imagenes a partir de la URL dada en el excel...
        try: 
            for i, foto_url in enumerate(columna):
                print("Entré al for..")
                print("y la i es: ", i)
                
                y = i
                image_id = tools.generaIDImagen(foto_url)
                print("Éste es el image id del q sacaremos el index row: ", image_id)
               
                #Nótese que index e i son distintos, donde index será la posición en donde debe ubicar la imagen.
                #Future: checar si por ende i quedó irrelevante.  

                #Actualiza la columna 'Name' con el nombre del archivo.
                dataframe.loc[i + nu_index, 'Name'] = image_id
                print("Imagen guardada en el dataframe...")
                print(f"Estoy en el for indice: {i + 1} de {len(columna)}.")
                print(f"En la row {i + nu_index} del gran total de {lote_total}.")
                #Importante: Si sabes que lo puedes hacer de una vez, puedes quitarle ese sleep para que lo haga superrápido.
                #Si en cambio sabes que habrá interrupciones, dejalo así, porque le das tiempo a si poder guardar el excel.
                
                #Guardaremos en excel cada 200 imagenes.
                #Future: Que la frecuencia de guardado se defina en globales.
                #Futue: Hay un bug que hace que ésto se ejecute al principio del recorrido, antes de llegar a los 200, corrige.
                if i % 200 == 0:
                    tools.df2Excel(dataframe, configuracion.sesion + '.xlsx')
                    print("Se guardará el excel cada 200 imagenes.")
                    
        except Exception as e:
                print(f"Excepción: - {e}.")
                print("Es probable que el archivo de excel esté abierto, ciérralo antes de proceder y oprime una tecla.")
                print("i = ", i)
                y = i
                input("Presiona cualquier tecla para continuar: ")
                print(f"Excepción: - {e}, guardaremos el excel hasta donde iba. Reinicia el proceso, continuará donde te quedaste.")
                #FUTURE: Haz que la excepción de excel abierto viva una sola vez, dentro de la función df2Excel.
                tools.df2Excel(dataframe, configuracion.sesion + '.xlsx')

        except KeyboardInterrupt:
            print("KEYBOARD: Interrumpiste el proceso, guardaré el dataframe en el excel, hasta donde ibamos.")
            #Como interrumpimos a proposito, no queremos que continúe el while, rebásalo.
            y = y + lote_procesar + 1
            print("Modifiqué y, y ahora vale: ", y)
            tools.df2Excel(dataframe, configuracion.sesion + '.xlsx')
            #Future: Incluir que con ESC cortemos el proceso.
        
    print("Terminó el ciclo que recorre las URLs, último guardado de excel...")
    tools.df2Excel(dataframe, configuracion.sesion + '.xlsx')
    
    return dataframe
    
def descargaImagenes(sesion):
    #Future: Estandariza todos los pydocs de cada función.
    """
    Recorre cada imagen obteniendo su nombre y guardándolo en el dataframe.
    Posteriormente obteniendo cada una de esas imagenes.

    Parameters:
    columna_fotos(dataframe): El objeto que contiene la columna con los nombres de las fotos.
    dataframe(dataframe): El dataframe final.

    Raises:

    Returns:
    dataframe:Regresa dataframe que se usará a través de los procesos.
    """

    #Future: Incluye el while de crear Excel aquí en descargaImagenes.

    #Define si ya existe el archivo de excel o se está completando un proceso previamente iniciado.
    if os.path.exists(globales.excel_results_path + configuracion.sesion + '.xlsx'):
        #Primero extraemos el dataframe:
        dataframe = pd.read_excel(globales.excel_results_path + configuracion.sesion + '.xlsx')
        print("Existe un archivo para éste lote e imprimiremos su dataframe...")
        print(dataframe)
        
        #Si ya existía traera Nan en los vacios.
        print("El archivo ya existía y estoy checando sus Nans en Download Status.")
        
        # Filtra las filas donde 'Download Status' es igual a 'Success'
        por_procesar = dataframe[dataframe['Download Status'].isna()]

        #Future: Que en el futuro haga un chequeo al directorio para ver si es necesario ya no bajarlas.
   
    else:
        #Crea el dataframe donde se registrarán los atributos y las difusiones con los campos necesarios.
        #Future: Si siempre creará el excel en éste punto, ve si sacas el punto B del 1.-CicloInicial.
        dataframe = creaExcel(configuracion.sesion + '.xlsx')
        #Si no existía traera ' '.
        print("Se creó el archivo de excel necesario y éste es su dataframe: ")
        print(dataframe)
        por_procesar = dataframe[dataframe['Download Status'] == '']
        print("Por procesar quedó así:")
        print(por_procesar)
                              
    cantidad_faltante = len(por_procesar)
    print(f"Por procesar tiene {cantidad_faltante} elementos.")
    
    # Crea un dataset 'columna_imagenes' a partir de la columna 'Nombre'
    columna_fotos = por_procesar[['Source', 'Name']]
    #df_imagenes_seleccionadas = df_images_ok[['Name', 'Source']]
    print("Imprime la columna que mide: ", len(columna_fotos))
    print(columna_fotos)

    contador = 1
    
    for index, row in columna_fotos.iterrows():
           
        Source = row['Source']
        Name = row['Name']

        #Attempt to download the image
        try:
            response = requests.get(Source)
            if response.status_code == 200:
                with open(f'imagenes/fuentes/{sesion}/{Name}', 'wb') as f:
                    f.write(response.content)
                download_status = 'Success'
                print(f"Image '{Name}' downloaded successfully. Índice: {index} de {len(columna_fotos)}.")
                
                
                ruta_total = f"imagenes\\fuentes\\{sesion}\\{Name}"
                print("Pretools: El resultado del SD fue exitoso, y su ruta total es/será: ", ruta_total)
                raiz_pc = os.getcwd()
                ruta_absoluta = os.path.join(raiz_pc, ruta_total)
                print("La ruta absoluta es desde donde le podré dar click desde el excel...", ruta_absoluta)
                
                dataframe.loc[index, 'Source Path'] = ruta_absoluta
                dataframe.loc[index, 'Download Status'] = download_status

                print("Listo, imagen guardada...")
                contador += 1

                #Guardaremos en excel cada 50 imagenes.
                #FUTURE: Definir la frecuencia de guardado en globales.
                if index % 50 == 0:
                    tools.df2Excel(dataframe, configuracion.sesion + '.xlsx')
                    print("50 imágenes más guardadas.")
                                    
            else:
                message = f"Error downloading image: {Source} (Status code: {response.status_code})"
                contador +=  1
                raise Exception(message)                

        except Exception as e:
            download_status = f"Error: {response.status_code}"
            dataframe.loc[index, 'Download Status'] = download_status
            print(f"Error downloading image: {Source} - {e}")

        except KeyboardInterrupt:
            print("KEYBOARD: Interrumpiste el proceso, guardaré el dataframe en el excel, hasta donde ibamos.")
            tools.df2Excel(dataframe, configuracion.sesion + '.xlsx')
        #Guarda excel de nuevo al acabar el for:
        tools.df2Excel(dataframe, configuracion.sesion + '.xlsx') 

def subeSources():
#Sube las imagenes source recién descargadas a mi propio servidor.

#Future: Importante...Agrega interrupción de teclado y guardados periódicos.

    sesion = configuracion.sesion
    base_url = globales.sources_url
    directorio_remoto = base_url + sesion

    print("Para empezar, éste es el directorio remoto:", directorio_remoto)
    
    #Conexión al servidor.
    ssh, sftp = servidor.conecta()  
    #Future: Revisar si sí es el excel de resultados con el que trabajaremos.
    excel = globales.excel_results_path + configuracion.sesion + '.xlsx'

    carpeta_remota = nodes.remote_sources
    #remote_sources = "/home/moibe/apps/holocards/sources/"
    print(f"La carpeta remota es: {carpeta_remota}")
    directorio_receptor = carpeta_remota + configuracion.sesion
    print(f"El directorio receptor será entonces: {directorio_receptor}.")

    try:       
        #Crea directorio
        print("Creando directorio, cuyo nombre será: ", directorio_receptor)
        #Si el directorio no existe, si lo está creando bien, checar después que problemas causa q ya exista.        
        sftp.mkdir(directorio_receptor)
        print("Directorio creado...")

    except Exception as e:
        # Mensaje de error
        print(f"Error al crear el directorio, probablemente ya existe: {e}")

    #Primero extraemos el dataframe:
    dataframe = pd.read_excel(excel) 
    #Parámetros: dataframe, columna_filtro, texto_filtro, columna_destino, columna_source 
    resultados = tools.getNotLoaded(dataframe, 'Download Status', 'Success', 'Source URL', 'Name')
    
    #Define ruta de la carpeta local donde se encuentran los sources.
    carpeta_local = globales.imagenes_folder_fuentes + sesion
    print("Ésta es la carpeta local: ", carpeta_local)
    
    print("Por entrar a cicloSubidor...")
    tools.cicloSubidor(sftp, dataframe, resultados, carpeta_local, directorio_receptor, directorio_remoto)

def directoriador(directorio):
    #FUTURE: Que los exceles iniciales residan en una carpeta exclusiva para eso.
    #FUTURE: Que el directoriador mande la carpeta hecha a un directorio específico.
    #FUTURE: Que source_excel y results_excel, estén en la misma carpeta.
    try:
        excel = globales.excel_results_path + directorio + ".xlsx"
        #Las imagenes tuvieron que haber sido subidas a la ruta correcta previamente.
        directory_address = "imagenes/fuentes/" + directorio
        print("El excel que usaremos es: ", excel)
        print("La ruta completa es: ", directory_address)
        
        #Future: Ve si lo estandarizas para que ésto también sea creado con pandas.
        workbook = load_workbook(excel)

    except FileNotFoundError:
        #Future: ¿Que hace ésto, crear un excel nuevo?
        workbook = Workbook()
    
    worksheet = workbook.active    
    worksheet.cell(row=1, column=1).value = "Name"
    worksheet.cell(row=1, column=2).value = "Download Status"
    worksheet.cell(row=1, column=3).value = "Take"
    worksheet.cell(row=1, column=4).value = "File"
    worksheet.cell(row=1, column=5).value = "Diffusion Status"
    
    row = 2  # Comenzar desde la fila 2 (después del encabezado)

    for filename in os.listdir(directory_address):
        
        # Agregar nombre de archivo en la primera columna
        worksheet.cell(row=row, column=1).value = filename

        # Agregar "Success" en la segunda columna
        worksheet.cell(row=row, column=2).value = "Success"

        row += 1

    workbook.save(excel)

def creaDirectorioResults(sesion):
    """
    Crea el directorio donde se recibirán los resultados en caso de no existir. El directorio llevará el nombre de la sesión + "results".

    Parameters:
    dataframe (dataframe): El dataframe en el que estuvimos trabajando.

    Returns:
    bool: True si se guardó el archivo correctamente.

    """
    #Future: Que la ruta venga de globales.
    results_dir = os.path.join('imagenes', 'resultados', sesion + "-results" )   
    
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)

#Mejor le pasas el objeto completo llamado creación y que de ahí lea cada uno de sus atributos.
#Future: Revisa si ya no se usa createColumns.
def createColumns(dataframe, amount, diccionario_atributos):
    #Future: Revisa si ésta ya está obsoleta.

    # Set of column names to repeat
    #column_names = ['DiffusionStatus', 'Take', 'Shot', 'Style', 'Hero', 'URL']
    column_names = list(diccionario_atributos)
    # Create a list of desired column order
    #desired_order = ['DiffusionStatus', 'Take', 'Shot', 'Style', 'Hero', 'URL']
    desired_order = ['DiffusionStatus'] + [name for name in column_names if name != 'DiffusionStatus'] + ['URL']

    # Sort column names based on desired order and repeat
    column_groups = [[f"{name}{i + 1}" for name in group] for i in range(amount) for group in zip(*sorted(zip(desired_order, column_names)))]
    
    
    # Flatten the nested list into a single list
    dynamic_columns = [item for sublist in column_groups for item in sublist]

    # Add dynamic columns to the DataFrame
    dataframe[dynamic_columns] = ''

    return dataframe

def preparaSamples(filename, samples):
    #Ésta función prepara el espacio para la cantidad de samples que quieres crear.
    #Creará una row adicional para cada sample que desees.

    #Primero extraemos el dataframe:
    dataframe = pd.read_excel(globales.excel_results_path + filename) 
    print("La cantidad de rows en el dataframe son: ", len(dataframe))  
    
    #Filtra las filas donde 'Download Status' es igual a 'Success' o 'From Archive'
    #Parámetros: dataframe, columnaAFiltrar, textoFiltro, textoFiltro2
    #Future, agregar filtros ilimitados con *kwargs.
    #IMPORTANTE: MUY IMPORTANTE. Si vas a llenar solo los From Archive, quite del filtro a Success.
    #En ocasiones posteriores no deberá pasar ésto porque se harán los dos al mismo tiempo.
    rowsFiltrados = tools.funcionFiltradora(dataframe, 'Download Status', 'Success', 'From Archive') 
    #IMPORTANTE: 'Form Archive' viene solo del hecho de haberse usado el recuperadorImagenes.py, creo que...
    #...ya no será necesario pero mantenerlo un rato.

    print("Ahora voy a imprimir rowsFiltrados, que tienen el tamaño:", len(rowsFiltrados))
    
    #Future: Hacer una función que seleccione las columna/s a usar.
    df_imagenes_seleccionadas = rowsFiltrados[['Name', 'Source', 'Source Path', 'Source URL']]
    cantidad_sampleos = samples * len(df_imagenes_seleccionadas)
    print("La cantidad de imagenes a trabajar será de : ", cantidad_sampleos)
        
    contador = 0
   
    try:

        #Crea las rows para sus samples
        for index, row in df_imagenes_seleccionadas.iterrows():
            imagen = row['Name']
            print("Imagen is: ", imagen)
            
            source = row['Source']
            print("Source is: ", source)
            
            source_path = row['Source Path']
            print("Source Path is: ", source_path)
            
            source_url = row['Source URL']
            #Ready!: Agregar a las repeticiones también la columna de Source Path.

            #Future: Que haga bien el conteo, porque mara 508 de 256.
            print(f"Procesadas {contador} de {cantidad_sampleos}.")
            
            nombre, extension = imagen.split(".")       

            #Cuando encuentra la imagen llena los datos de la primera incidencia.
            indice = tools.obtenIndexRow(dataframe, 'Name', imagen)
            dataframe.loc[indice, 'Take'] = 1
            dataframe.loc[indice, 'File'] = nombre + "-" + "t" + str(1) + "." + extension
            #FUTURE: Aquí causa el error de type de pandas, corrigelo antes de que quede deprecado.
            
            #AQUÍ VA IR LO QUE TENIAMOS DENTRO DEL FOR, que son el resto de los samples:
            if configuracion.excel_list is True:
                #En lugar de esas comillas vas a poner el source.
                #En excel: [Source, Source Path, Source URL, Name, Download Status, Take, File, Diffusion Status]
                lista = [source, source_path, source_url, imagen, 'Success', "take_placeholder", "filename", "", "", ""]  #adding a row
                #Designación de columnas a utilizar.
                a = 5 #Aquí sustituirá el índice 3 take_placeholder
                b = 6 
            else:         
                #Para imagenes de directorio.
                #FUTURE: Definir si agregaras campos de path local y url en modo directoriador.
                lista = [imagen, 'Success', "take_placeholder", "filename", "", "",""]  #adding a row
                #Designación de columnas a utilizar.
                a = 6 #Aquí sustituirá el índice 4 take_placeholder
                b = 7

            #Y luego crea tantas rows adicionales como samples fuera a haber.
            for i in range(samples - 1): 
                        
                # Replace the element at the index with the sustituto variable
                #Esto es el Take
                lista[a] = i + 2

                #Empieza desde el 2 porque ya hizo la 1.
                #Esto es el File
                filename = nombre + "-" + "t" + str(i+2) + "." + extension
                
                # Replace the element at the index with the sustituto variable
                lista[b] = filename
                
                tools.creaRow(dataframe, imagen, i + 2, filename, lista)

            contador = contador + 1
    
    except KeyboardInterrupt:
      print("KEYBOARD: Interrumpiste el proceso, guardaré el dataframe en el excel, hasta donde ibamos. Excel:", configuracion.filename)
      dataframe = dataframe.sort_values(['Name','Take'])
      tools.df2Excel(dataframe, configuracion.sesion + '.xlsx')        
    
    #Reordeno alfabéticamente.
    #El ordenamiento si es necesario, así es que también incluyelo en el Keyboard Interrupt.
    dataframe = dataframe.sort_values(['Name','Take'])
    tools.df2Excel(dataframe, configuracion.sesion + '.xlsx')
      
    #Future: Revisa, creo que no es necesario usar el return frame.
    #return dataframe